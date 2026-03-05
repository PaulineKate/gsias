#!/usr/bin/env python3
"""
update_excel.py  --  payroll Excel generator
"""

import sys, json, os

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Run: py -m pip install openpyxl")
    sys.exit(1)

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "payroll.xlsx")

# ── Layout constants ──────────────────────────────────────────────
COL_OFFSET  = 0      # No left-side padding column — tables start at column A
ROW_START   = 1      # First table begins at row 1
GAP_ROWS    = 19     # Blank rows between end of one table and start of the next
MAX_PERIODS = 13
HDR_ROWS    = 5

# ── Column indices (all start at A now that COL_OFFSET = 0) ──────
C_PERIOD  = 1  + COL_OFFSET   # A
C_DAYS    = 2  + COL_OFFSET   # B
C_RATE    = 3  + COL_OFFSET   # C
C_WAGE    = 4  + COL_OFFSET   # D
C_LBP     = 5  + COL_OFFSET   # E
C_PCONT   = 6  + COL_OFFSET   # F
C_PMPL    = 7  + COL_OFFSET   # G
C_SSS     = 8  + COL_OFFSET   # H
C_LATE    = 9  + COL_OFFSET   # I
C_NURSERY = 10 + COL_OFFSET   # J
C_AMTDUE  = 11 + COL_OFFSET   # K

MONTH_NAMES = ["jan", "feb", "mar", "apr", "may", "jun",
               "jul", "aug", "sep", "oct", "nov", "dec"]

# ── Style helpers ─────────────────────────────────────────────────
def T(style="thin"): return Side(style=style)
def N():             return Side(style=None)
def ba():            return Border(left=T(), right=T(), top=T(), bottom=T())
def Fnt(bold=False, size=9): return Font(name="Arial", bold=bold, size=size)
def AL(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def rh(ws, r, h): ws.row_dimensions[r].height = h

def col_letter(col_idx):
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result

LD = col_letter(C_DAYS)
LR = col_letter(C_RATE)
LG = col_letter(C_WAGE)
LE = col_letter(C_LBP)
LJ = col_letter(C_NURSERY)

# ── Period text formatter ─────────────────────────────────────────
def format_period_text(text):
    if not text:
        return ""
    if " - " in text:
        text = text.replace(" - ", " -\n")
    elif " & " in text:
        text = text.replace(" & ", " &\n")
    return "\n" + text + "\n"

# ── Month helpers ─────────────────────────────────────────────────
def extract_month(period_label):
    label = period_label.strip().lower()
    for i, m in enumerate(MONTH_NAMES):
        if label.startswith(m):
            return i
    return -1

def month_sort_key(entry):
    p = entry["period"].lower()
    m = extract_month(p)
    sub_sort = 0
    if m == 11:
        if "16-31" in p:   sub_sort = 1
        elif "1-31" in p:  sub_sort = 2
        else:              sub_sort = 3
    return (m if m >= 0 else 99, sub_sort, p)

# ── Sheet / column setup ─────────────────────────────────────────
def set_col_widths(ws):
    widths = {
        C_PERIOD: 12, C_DAYS: 8,  C_RATE: 10, C_WAGE: 12,
        C_LBP:    9,  C_PCONT: 10, C_PMPL: 9,  C_SSS:  10,
        C_LATE:   10, C_NURSERY: 11, C_AMTDUE: 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col_letter(col)].width = w

def get_sheet(wb, year_str):
    if year_str in wb.sheetnames:
        return wb[year_str]
    ws = wb.create_sheet(title=year_str)
    set_col_widths(ws)
    return ws

# ── Block finder ─────────────────────────────────────────────────
def find_block(ws, name):
    """
    Find the starting row of an existing person block.
    The block header row has "NAME" in C_PERIOD and the person's name in C_DAYS.
    """
    name_upper = name.strip().upper()
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if not isinstance(cell.value, str):
                continue
            if cell.value.strip().upper() == name_upper:
                # The name cell is at C_DAYS; check that C_PERIOD on the same row == "NAME"
                left = ws.cell(row=cell.row, column=C_PERIOD)
                if isinstance(left.value, str) and left.value.strip().upper() == "NAME":
                    return cell.row
    return None

# ── Next free row calculation ────────────────────────────────────
def block_height():
    """Total rows consumed by one complete block (headers + data rows)."""
    return HDR_ROWS + MAX_PERIODS  # = 5 + 13 = 18

def next_free_row(ws):
    """
    Return the row where the next new block should start.

    Rules:
    - If the sheet is empty, start at ROW_START.
    - Otherwise find the last occupied row across all data columns,
      then add GAP_ROWS (19) blank rows after it.
    """
    last_used = 0
    for r in range(ws.max_row, 0, -1):
        vals = [ws.cell(row=r, column=c).value
                for c in range(C_PERIOD, C_AMTDUE + 1)]
        if any(v not in (None, "") for v in vals):
            last_used = r
            break

    if last_used == 0:
        return ROW_START

    # last_used is the last row with content; next block starts GAP_ROWS later
    return last_used + GAP_ROWS + 1

# ── Block writer ─────────────────────────────────────────────────
def write_block(ws, start_row, name, designation, year):
    r = start_row
    s = T()

    # Row 1 of block — Name / Designation header
    rh(ws, r, 30)
    c = ws.cell(row=r, column=C_PERIOD, value="Name")
    c.border    = ba()
    c.alignment = AL("left", "top")

    ws.merge_cells(start_row=r, start_column=C_DAYS,
                   end_row=r, end_column=C_WAGE)
    for col in range(C_DAYS, C_WAGE + 1):
        ws.cell(row=r, column=col).border = ba()
    c = ws.cell(row=r, column=C_DAYS, value=name.upper())
    c.font      = Font(name="Arial", bold=True, size=10)
    c.alignment = AL("center", "center")

    c = ws.cell(row=r, column=C_LBP, value="Position")
    c.border    = ba()
    c.alignment = AL("left", "top")

    # Designation merges C_PCONT→C_LATE (wide cell), then C_NURSERY and C_AMTDUE
    # are separate bordered empty cells — this creates the divider line aligned
    # with the Late Deduction / Nursery Product column boundary below.
    ws.merge_cells(start_row=r, start_column=C_PCONT,
                   end_row=r, end_column=C_LATE)
    for col in range(C_PCONT, C_LATE + 1):
        left_side  = T() if col == C_PCONT else N()
        right_side = T() if col == C_LATE  else N()
        ws.cell(row=r, column=col).border = Border(
            left=left_side, right=right_side, top=T(), bottom=T()
        )
    c = ws.cell(row=r, column=C_PCONT, value=designation)
    c.font      = Fnt()
    c.alignment = AL("center", "center", True)
    # Separate cells to the right — no value, just borders to complete the row
    ws.cell(row=r, column=C_NURSERY).border = Border(left=T(), right=N(), top=T(), bottom=T())
    ws.cell(row=r, column=C_AMTDUE).border  = Border(left=N(), right=T(), top=T(), bottom=T())

    r += 1

    # Rows 2-3 — Annual Salary / Effective Date
    for i in range(2):
        rh(ws, r, 18)
        ws.merge_cells(start_row=r, start_column=C_PERIOD, end_row=r, end_column=C_WAGE)
        ws.merge_cells(start_row=r, start_column=C_LBP,    end_row=r, end_column=C_AMTDUE)
        for col in range(C_PERIOD, C_AMTDUE + 1):
            lb = s if col in (C_PERIOD, C_LBP)   else N()
            rb = s if col in (C_WAGE,   C_AMTDUE) else N()
            bb = s if i == 1 else N()
            ws.cell(row=r, column=col).border = Border(
                left=lb, right=rb, top=N(), bottom=bb)
        if i == 0:
            ws.cell(row=r, column=C_PERIOD, value="Annual Salary").font  = Fnt()
            ws.cell(row=r, column=C_LBP,    value="Effective Date of Appointment").font = Fnt()
        else:
            ws.cell(row=r, column=C_PERIOD, value=str(year)).font = Fnt()
        r += 1

    # Rows 4-5 — Column headers (spanning two rows)
    r4, r5 = r, r + 1
    rh(ws, r4, 22)
    rh(ws, r5, 18)

    span_cols = {
        C_PERIOD:  "PERIOD\nCOVERED",
        C_DAYS:    "NO. OF\nDAYS",
        C_RATE:    "RATE PER\nDAY",
        C_WAGE:    "TOTAL\nWAGE",
        C_LBP:     "LBP",
        C_SSS:     "SSS\nCONT.",
        C_LATE:    "Late\nDeduction",
        C_NURSERY: "NURSERY\nPRODUCT",
        C_AMTDUE:  "AMOUNT\nDUE",
    }
    for col, label in span_cols.items():
        ws.merge_cells(start_row=r4, start_column=col,
                       end_row=r5,   end_column=col)
        c = ws.cell(row=r4, column=col, value=label)
        c.font      = Fnt(True, 8)
        c.alignment = AL("center", wrap=True)
        c.border    = ba()
        ws.cell(row=r5, column=col).border = ba()

    ws.merge_cells(start_row=r4, start_column=C_PCONT,
                   end_row=r4,   end_column=C_PMPL)
    c = ws.cell(row=r4, column=C_PCONT, value="Pag-ibig")
    c.font      = Fnt(True, 8)
    c.alignment = AL("center")
    c.border    = Border(left=s, right=s, top=s)

    for col, lbl in [(C_PCONT, "Cont."), (C_PMPL, "MPL")]:
        c = ws.cell(row=r5, column=col, value=lbl)
        c.font      = Fnt(True, 8)
        c.alignment = AL("center")
        c.border    = ba()

    r += 2  # skip r4 and r5

    # Data rows (MAX_PERIODS empty slots)
    for _ in range(MAX_PERIODS):
        _write_empty_period_row(ws, r)
        r += 1

# ── Data row helpers ─────────────────────────────────────────────
def _write_empty_period_row(ws, r):
    for col in range(C_PERIOD, C_AMTDUE + 1):
        c = ws.cell(row=r, column=col)
        c.font   = Fnt()
        c.border = ba()
        if col == C_DAYS:
            c.number_format = "0.0"
            c.alignment     = AL("center", "center")
        elif col == C_PERIOD:
            c.alignment = AL("left", "center", wrap=True)
        else:
            c.number_format = "#,##0.00"
            c.alignment     = AL("right", "center")
        if col == C_WAGE:
            c.value = f"=IFERROR({LD}{r}*{LR}{r},0)"
            c.font  = Fnt(True)
        if col == C_AMTDUE:
            c.value = f"=IFERROR({LG}{r}-SUM({LE}{r}:{LJ}{r}),0)"
            c.font  = Fnt(True)
    rh(ws, r, 18)

def _write_data_row(ws, r, period, days, rate, lbp, pcont, pmpl, sss, late, nursery):
    final_text = format_period_text(period)
    _write_empty_period_row(ws, r)
    ws.cell(row=r, column=C_PERIOD,  value=final_text)
    ws.cell(row=r, column=C_DAYS,    value=days)
    ws.cell(row=r, column=C_RATE,    value=rate)
    ws.cell(row=r, column=C_LBP,     value=lbp)
    ws.cell(row=r, column=C_PCONT,   value=pcont)
    ws.cell(row=r, column=C_PMPL,    value=pmpl)
    ws.cell(row=r, column=C_SSS,     value=sss)
    ws.cell(row=r, column=C_LATE,    value=late)
    ws.cell(row=r, column=C_NURSERY, value=nursery)
    # Taller row when the padded text wraps to multiple visible lines
    rh(ws, r, 35 if final_text.count("\n") >= 3 else 25)

# ── Period read / fill ────────────────────────────────────────────
def read_period_rows(ws, block_start_row):
    first_data_row = block_start_row + HDR_ROWS
    rows = []
    for i in range(MAX_PERIODS):
        r   = first_data_row + i
        val = ws.cell(row=r, column=C_PERIOD).value
        if not val:
            continue
        clean = str(val).strip().replace("\n", " ").replace("  ", " ").strip()
        rows.append({
            "period":  clean,
            "days":    ws.cell(row=r, column=C_DAYS).value,
            "rate":    ws.cell(row=r, column=C_RATE).value,
            "lbp":     ws.cell(row=r, column=C_LBP).value,
            "pcont":   ws.cell(row=r, column=C_PCONT).value,
            "pmpl":    ws.cell(row=r, column=C_PMPL).value,
            "sss":     ws.cell(row=r, column=C_SSS).value,
            "late":    ws.cell(row=r, column=C_LATE).value,
            "nursery": ws.cell(row=r, column=C_NURSERY).value,
        })
    return rows

def fill_period(ws, block_row, period, record, rate):
    existing  = read_period_rows(ws, block_row)
    new_entry = {
        "period":  period,
        "days":    record.get("num_days"),
        "rate":    rate,
        "lbp":     record.get("lbp"),
        "pcont":   record.get("pagibig_cont"),
        "pmpl":    record.get("pagibig_mpl"),
        "sss":     record.get("sss_cont"),
        "late":    record.get("late_deduction"),
        "nursery": record.get("nursery_prod"),
    }
    replaced = False
    merged   = []
    for entry in existing:
        if entry["period"].strip().lower() == period.strip().lower():
            merged.append(new_entry)
            replaced = True
        else:
            merged.append(entry)
    if not replaced:
        merged.append(new_entry)
    merged.sort(key=month_sort_key)
    if len(merged) > MAX_PERIODS:
        merged = merged[:MAX_PERIODS]

    first_data_row = block_row + HDR_ROWS
    for i in range(MAX_PERIODS):
        r = first_data_row + i
        if i < len(merged):
            e = merged[i]
            _write_data_row(ws, r, e["period"], e["days"], e["rate"],
                            e["lbp"], e["pcont"], e["pmpl"],
                            e["sss"], e["late"], e["nursery"])
        else:
            _write_empty_period_row(ws, r)

# ── Entry point ───────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    year        = str(data.get("year", ""))
    name        = data.get("name", "").strip().upper()
    designation = data.get("designation", "").strip()
    period      = data.get("period_covered", "").strip()
    rate        = float(data.get("rate", 0))

    def v(k):
        try:
            return float(data.get(k, 0)) or None
        except Exception:
            return None

    record = {
        "num_days":       v("num_days"),
        "lbp":            v("lbp"),
        "pagibig_cont":   v("pagibig_cont"),
        "pagibig_mpl":    v("pagibig_mpl"),
        "sss_cont":       v("sss_cont"),
        "late_deduction": v("late_deduction"),
        "nursery_prod":   v("nursery_prod"),
    }

    wb = load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else Workbook()
    if not os.path.exists(EXCEL_PATH):
        for sn in wb.sheetnames:
            del wb[sn]

    ws        = get_sheet(wb, year)
    block_row = find_block(ws, name)

    if block_row is None:
        block_row = next_free_row(ws)
        write_block(ws, block_row, name, designation, year)

    fill_period(ws, block_row, period, record, rate)
    wb.save(EXCEL_PATH)

if __name__ == "__main__":
    main()